import tkinter as tk
from tkinter import messagebox, filedialog
import pandas as pd
from num2words import num2words
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# --- SIP Function ---
def calculate_sip(principal, annual_rate, years):
    monthly_rate = annual_rate / 12
    months = years * 12
    data = []
    opening = 0
    for month in range(1, months + 1):
        interest = (opening + principal) * monthly_rate
        closing = opening + principal + interest
        data.append({
            'Month': month,
            'Opening Balance (₹)': round(opening, 2),
            'Monthly Investment (₹)': principal,
            'Interest Earned (₹)': round(interest, 2),
            'Closing Balance (₹)': round(closing, 2)
        })
        opening = closing
    df = pd.DataFrame(data)
    return df, principal * months, round(opening, 2)

# --- SWP Function ---
def calculate_swp(initial_corpus, withdrawal, annual_rate, years):
    monthly_rate = annual_rate / 12
    months = years * 12
    data = []
    opening = initial_corpus
    for month in range(1, months + 1):
        interest = opening * monthly_rate
        closing = opening + interest - withdrawal
        data.append({
            'Month': month,
            'Opening Balance (₹)': round(opening, 2),
            'Interest Earned (₹)': round(interest, 2),
            'Monthly Withdrawal (₹)': withdrawal,
            'Closing Balance (₹)': round(closing, 2)
        })
        if closing <= 0:
            break
        opening = closing
    df = pd.DataFrame(data)
    return df, round(opening, 2)

# --- Generate Excel ---
def generate_excel():
    try:
        sip_amt = float(entry_sip_amt.get())
        sip_years = int(entry_sip_years.get())
        sip_return = float(entry_sip_return.get()) / 100

        swp_amt = float(entry_swp_amt.get())
        swp_years = int(entry_swp_years.get())
        swp_return = float(entry_swp_return.get()) / 100

        # SIP Calculation
        sip_df, sip_invested, sip_final = calculate_sip(sip_amt, sip_return, sip_years)
        sip_gain = sip_final - sip_invested
        sip_final_words = num2words(sip_final, lang='en_IN', to='currency').replace("euro", "rupees").replace("cents", "paise")

        # SWP Calculation
        swp_df, swp_remaining = calculate_swp(sip_final, swp_amt, swp_return, swp_years)
        swp_remaining_words = num2words(swp_remaining, lang='en_IN', to='currency').replace("euro", "rupees").replace("cents", "paise")

        # Summary
        summary_df = pd.DataFrame({
            "Metric": [
                "SIP Monthly Investment", "SIP Duration", "SIP Annual Return",
                "Total SIP Invested", "Final Corpus (SIP)", "Corpus in Words",
                "SIP Returns (Corpus - Invested)",
                "SWP Monthly Withdrawal", "SWP Duration", "SWP Annual Return",
                "Remaining Corpus after SWP", "Remaining Corpus in Words"
            ],
            "Value": [
                f"₹{sip_amt:,.2f}", f"{sip_years} years", f"{sip_return*100:.2f}%",
                f"₹{sip_invested:,.2f}", f"₹{sip_final:,.2f}", sip_final_words,
                f"₹{sip_gain:,.2f}", f"₹{swp_amt:,.2f}", f"{swp_years} years",
                f"{swp_return*100:.2f}%", f"₹{swp_remaining:,.2f}", swp_remaining_words
            ]
        })

        # Save Excel
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel File"
        )
        if not file_path:
            return

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "SIP Calculation"
        for row in dataframe_to_rows(sip_df, index=False, header=True):
            ws1.append(row)

        chart = LineChart()
        chart.title = "SIP Closing Balance"
        chart.y_axis.title = "₹"
        chart.x_axis.title = "Month"
        data = Reference(ws1, min_col=5, min_row=1, max_row=ws1.max_row)
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws1.add_chart(chart, f"A{ws1.max_row + 3}")

        ws2 = wb.create_sheet("SWP Plan")
        for row in dataframe_to_rows(swp_df, index=False, header=True):
            ws2.append(row)

        ws3 = wb.create_sheet("Summary")
        for row in dataframe_to_rows(summary_df, index=False, header=True):
            ws3.append(row)

        wb.save(file_path)
        messagebox.showinfo("Success", f"Excel file saved:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# --- Tkinter UI ---
root = tk.Tk()
root.title("SIP + SWP Excel Generator")

tk.Label(root, text="SIP Monthly Investment (₹):").grid(row=0, column=0, sticky="e")
entry_sip_amt = tk.Entry(root)
entry_sip_amt.grid(row=0, column=1)

tk.Label(root, text="SIP Duration (years):").grid(row=1, column=0, sticky="e")
entry_sip_years = tk.Entry(root)
entry_sip_years.grid(row=1, column=1)

tk.Label(root, text="SIP Annual Return (%):").grid(row=2, column=0, sticky="e")
entry_sip_return = tk.Entry(root)
entry_sip_return.grid(row=2, column=1)

tk.Label(root, text="SWP Monthly Withdrawal (₹):").grid(row=3, column=0, sticky="e")
entry_swp_amt = tk.Entry(root)
entry_swp_amt.grid(row=3, column=1)

tk.Label(root, text="SWP Duration (years):").grid(row=4, column=0, sticky="e")
entry_swp_years = tk.Entry(root)
entry_swp_years.grid(row=4, column=1)

tk.Label(root, text="SWP Annual Return (%):").grid(row=5, column=0, sticky="e")
entry_swp_return = tk.Entry(root)
entry_swp_return.grid(row=5, column=1)

tk.Button(root, text="Generate Excel", command=generate_excel, bg="green", fg="white").grid(row=6, column=0, columnspan=2, pady=10)

root.mainloop()
